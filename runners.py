"""Runner classes for capture, extraction, and unified crawl.

JS strings: build_seo_js() and build_extraction_js() return JavaScript as
Python f-strings/raw-strings. These are injected via sb.cdp.evaluate() and
run in the browser context. No template engine — just string concatenation
with field lists joined by semicolons.
"""

from __future__ import annotations

import csv
import io
import json
import random
import re
import threading
import time
import zipfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from typing import Callable, Optional

import img2pdf
from PIL import Image
from seleniumbase import SB

from extraction import (
    build_seo_js,
    extract_from_page,
    get_standard_seo_fields,
    parse_seo_fields,
)
from importers import is_valid_url
from page_capture import PageCapture

HERE = Path(__file__).resolve().parent
HISTORY_FILE = HERE / ".run_history.json"


def slugify(url: str) -> str:
    base = re.sub(r"^https?://", "", url.lower())
    return re.sub(r"[^a-z0-9]+", "_", base).strip("_")


def parse_seo_payload(raw: str, seo_fields: list[dict] | None = None) -> dict:
    """Parse raw JSON from SEO JS into a flat dict.

    When seo_fields is provided, uses the configurable field parser (parse_seo_fields).
    When seo_fields is None, falls back to a legacy hardcoded mapping that matches
    the original _seo_js() output format exactly. This fallback exists for backwards
    compatibility with callers that don't pass seo_fields (e.g. tests, run_capture.py).
    """
    if seo_fields is not None:
        return parse_seo_fields(raw, seo_fields)
    # Legacy fallback — matches original _seo_js() output exactly
    payload = json.loads(raw or "{}")
    return {
        "title": payload.get("title", ""),
        "title_len": len(payload.get("title", "")),
        "meta_description": payload.get("metaDesc", ""),
        "meta_desc_len": len(payload.get("metaDesc", "")),
        "canonical": payload.get("canonical", ""),
        "robots_meta": payload.get("robotsMeta", ""),
        "h1": payload.get("h1", ""),
        "h2s": payload.get("h2s", ""),
        "h3s": payload.get("h3s", ""),
        "og_title": payload.get("ogTitle", ""),
        "og_description": payload.get("ogDesc", ""),
        "og_image": payload.get("ogImage", ""),
        "schema_types": payload.get("schemaTypes", ""),
        "word_count": payload.get("wordCount", 0),
        "internal_links": payload.get("internal", 0),
        "external_links": payload.get("external", 0),
        "images_missing_alt": payload.get("imagesMissingAlt", 0),
    }


def write_results_csv(results: list[dict], csv_path: Path) -> None:
    """Write a list of result dicts to a CSV file."""
    if not results:
        return
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    all_keys = list(dict.fromkeys(k for row in results for k in row))
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=all_keys, extrasaction="ignore")
        w.writeheader()
        w.writerows(results)


def write_results_excel(results: list[dict], xlsx_path: Path) -> None:
    """Write SEO results to an Excel workbook with multiple sheets."""
    if not results:
        return
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    # ── Sheet 1: SEO data (flat fields only) ──
    flat_keys = [
        "url", "status", "status_code", "title", "title_len",
        "meta_description", "meta_desc_len", "canonical", "canonical_chain",
        "robots_meta", "h1", "h2s", "h3s",
        "og_title", "og_description", "og_image", "og_type",
        "twitter_card", "twitter_title", "twitter_description", "twitter_image",
        "schema_types", "word_count", "internal_links", "external_links",
        "images_missing_alt", "images_total", "images_no_lazy",
        "iframe_count", "form_count", "external_nofollow",
        "html_lang", "meta_viewport", "meta_charset", "hreflang",
        "depth", "redirect_chain", "page_size", "response_time",
        "redirected_url", "redirected_status_code",
        "internal_inlinks", "content_type", "boilerplate_ratio",
    ]
    ws1 = wb.active or wb.create_sheet("SEO Data")
    ws1.title = "SEO Data"
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")
    for col, key in enumerate(flat_keys, 1):
        cell = ws1.cell(row=1, column=col, value=key)
        cell.font = header_font_white
        cell.fill = header_fill
    for row_idx, row_data in enumerate(results, 2):
        for col_idx, key in enumerate(flat_keys, 1):
            ws1.cell(row=row_idx, column=col_idx, value=row_data.get(key, ""))

    # Auto-adjust column widths
    for col in ws1.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        first_cell = col[0]
        col_letter = first_cell.column_letter  # type: ignore[attr-defined]
        ws1.column_dimensions[col_letter].width = min(max_len + 2, 60)

    # ── Sheet 2: Issues summary (if build_issues found) ──
    ws2 = wb.create_sheet("Issues")
    try:
        from analysis import analyze_results
        issues = analyze_results(results)
        ws2.cell(row=1, column=1, value="Category").font = header_font
        ws2.cell(row=1, column=2, value="Severity").font = header_font
        ws2.cell(row=1, column=3, value="Issue").font = header_font
        ws2.cell(row=1, column=4, value="Affected URLs").font = header_font
        for i, issue in enumerate(issues, 2):
            ws2.cell(row=i, column=1, value=issue.category)
            ws2.cell(row=i, column=2, value=issue.severity)
            ws2.cell(row=i, column=3, value=issue.name)
            ws2.cell(row=i, column=4, value=issue.count)
        ws2.column_dimensions["A"].width = 18
        ws2.column_dimensions["B"].width = 14
        ws2.column_dimensions["C"].width = 30
        ws2.column_dimensions["D"].width = 14
    except Exception:
        ws2.cell(row=1, column=1, value="Issue analysis not available")

    wb.save(xlsx_path)


def write_results_json(results: list[dict], json_path: Path) -> None:
    """Write results as a JSON file."""
    if not results:
        return
    import json
    json_path.parent.mkdir(parents=True, exist_ok=True)
    data = {
        "exported_at": datetime.now().isoformat(),
        "total": len(results),
        "ok": sum(1 for r in results if r.get("status") == "ok"),
        "fail": sum(1 for r in results if r.get("status") != "ok"),
        "results": results,
    }
    with json_path.open("w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, default=str)


def _compute_crawl_statistics(results: list[dict]) -> dict:
    """Compute aggregate crawl statistics from results."""
    ok = [r for r in results if r.get("status") == "ok"]
    total = len(results)
    return {
        "total": total,
        "ok": len(ok),
        "fail": total - len(ok),
        "by_status_code": {
            str(sc): sum(1 for r in results if r.get("status_code") == sc)
            for sc in sorted({r.get("status_code", 0) for r in results})
        },
        "by_content_type": {
            ct: sum(1 for r in ok if r.get("content_type") == ct)
            for ct in sorted({r.get("content_type", "") for r in ok if r.get("content_type")})
        },
        "by_depth": {
            str(d): sum(1 for r in ok if r.get("depth") == d)
            for d in sorted({r.get("depth", 0) for r in ok})
        },
        "total_internal_links": sum(r.get("internal_links", 0) for r in ok),
        "total_external_links": sum(r.get("external_links", 0) for r in ok),
        "avg_word_count": round(sum(r.get("word_count", 0) or 0 for r in ok) / max(len(ok), 1), 1),
        "avg_page_size": round(sum(r.get("page_size", 0) or 0 for r in ok) / max(len(ok), 1), 1),
        "avg_response_time": round(sum(r.get("response_time", 0) or 0 for r in ok) / max(len(ok), 1), 1),
        "pages_with_broken_links": sum(1 for r in ok if r.get("broken_links")),
        "pages_with_canonical_chain": sum(1 for r in ok if r.get("canonical_chain")),
        "pages_with_redirects": sum(1 for r in ok if r.get("redirected_url")),
    }


def build_runtime_config(CONFIG: dict, viewport: dict, stabilization_ms: int) -> dict:
    """Build the runtime config dict from global CONFIG and per-run overrides."""
    return {
        "viewport": viewport,
        "timing": {**CONFIG["timing"], "stabilization_ms": stabilization_ms},
        "hide": CONFIG.get("hide", {}),
        "hide_visibility": CONFIG.get("hide_visibility", {}),
        "crawl4ai": CONFIG.get("crawl4ai", {}),
    }


def _compute_internal_inlinks(results: list[dict]) -> None:
    """Compute internal inlink counts by building a link graph from stored link details.

    Each result row should have internal_links_detail containing {href, text, rel}
    for every outbound internal link found on that page. This function inverts
    that mapping to count how many pages link TO each URL.
    """
    from urllib.parse import urlparse, urlunparse

    inlink_map: dict[str, set[str]] = {}

    def _normalize(u: str) -> str:
        parsed = urlparse(u)
        return urlunparse((parsed.scheme, parsed.netloc, parsed.path.rstrip("/"), "", "", ""))

    for row in results:
        if row.get("status") != "ok":
            continue
        source_url = _normalize(row.get("url", ""))
        if not source_url:
            continue
        for link in row.get("internal_links_detail", []):
            target = _normalize(link.get("href", ""))
            if target and target != source_url:
                inlink_map.setdefault(target, set()).add(source_url)

    for row in results:
        url = _normalize(row.get("url", ""))
        if url and url in inlink_map:
            row["internal_inlinks"] = len(inlink_map[url])
        else:
            row["internal_inlinks"] = 0


def _analyze_broken_links(results: list[dict]) -> list[dict]:
    """Scan all outlinks found during crawl and check for 4xx/5xx responses.

    Returns a list of broken link entries: {source_url, target_url, status_code}.
    """
    import requests

    all_outlinks: set[str] = set()
    source_map: dict[str, set[str]] = {}
    for row in results:
        src = row.get("url", "")
        if not src:
            continue
        for link in row.get("internal_links_detail", []):
            href = link.get("href", "")
            if href and href.startswith("http"):
                all_outlinks.add(href)
                source_map.setdefault(href, set()).add(src)
        for link in row.get("external_links_detail", []):
            href = link.get("href", "")
            if href and href.startswith("http"):
                all_outlinks.add(href)
                source_map.setdefault(href, set()).add(src)

    broken: list[dict] = []
    session = requests.Session()
    session.headers.update({"User-Agent": "Mozilla/5.0 PageCapture/1.0 broken-link-check"})
    for href in all_outlinks:
        try:
            resp = session.head(href, timeout=10, allow_redirects=True)
            if resp.status_code >= 400:
                for src_url in source_map.get(href, set()):
                    broken.append({
                        "source_url": src_url,
                        "target_url": href,
                        "status_code": resp.status_code,
                    })
        except requests.RequestException:
            for src_url in source_map.get(href, set()):
                broken.append({
                    "source_url": src_url,
                    "target_url": href,
                    "status_code": 0,
                })

    session.close()
    return broken


def _resolve_canonical_chains(results: list[dict]) -> None:
    """Follow canonical chains to their final destination for each result row."""
    import requests

    url_map = {r.get("url", ""): r for r in results if r.get("url")}
    session = requests.Session()
    session.headers.update({"User-Agent": "Mozilla/5.0 PageCapture/1.0 canonical-resolver"})
    for row in results:
        canon = row.get("canonical", "")
        if not canon or canon == row.get("url"):
            continue
        chain = [canon]
        visited: set[str] = set()
        current = canon
        while current and current not in visited:
            visited.add(current)
            if current in url_map:
                next_canon = url_map[current].get("canonical", "")
                if next_canon and next_canon != current:
                    chain.append(next_canon)
                    current = next_canon
                else:
                    break
            else:
                try:
                    resp = session.head(current, timeout=10, allow_redirects=True)
                    if resp.status_code < 400:
                        final_url = resp.url
                        if final_url != current:
                            chain.append(final_url)
                    break
                except requests.RequestException:
                    break
        row["canonical_chain"] = " -> ".join(chain)
    session.close()


def build_zip(results: list[dict], output_dir: Path) -> bytes:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for r in results:
            for key in ("file", "pdf"):
                raw = r.get(key, "")
                if raw:
                    p = Path(raw)
                    if p.is_file():
                        zf.write(p, arcname=p.name)
    return buf.getvalue()


_PDF_DPI = 150


def png_to_pdf(png_path: Path, pdf_path: Path) -> None:
    """Convert a PNG to a single-page PDF using img2pdf with correct DPI."""
    dpi = _PDF_DPI
    with Image.open(png_path) as im:
        raw_dpi = im.info.get("dpi")
        if raw_dpi and raw_dpi[0] > 0 and raw_dpi[1] > 0:
            dpi = (int(raw_dpi[0]), int(raw_dpi[1]))
    pdf_path.parent.mkdir(parents=True, exist_ok=True)
    pdf_bytes = img2pdf.convert(str(png_path), dpi=dpi)
    if pdf_bytes:
        with open(pdf_path, "wb") as f:
            f.write(pdf_bytes)


def load_history() -> list[dict]:
    if HISTORY_FILE.exists():
        try:
            with HISTORY_FILE.open(encoding="utf-8") as f:
                return json.load(f)
        except (json.JSONDecodeError, OSError):
            return []
    return []


def save_history(entry: dict) -> None:
    history = load_history()
    history.insert(0, entry)
    if len(history) > 50:
        history = history[:50]
    with HISTORY_FILE.open("w", encoding="utf-8") as f:
        json.dump(history, f, indent=2)


def delete_history_entry(index: int) -> None:
    history = load_history()
    if 0 <= index < len(history):
        history.pop(index)
        with HISTORY_FILE.open("w", encoding="utf-8") as f:
            json.dump(history, f, indent=2)


def get_results(entry: dict) -> dict[str, list[dict]]:
    """Normalize history entry results into a collector-keyed dict.

    UnifiedRunner saves ``results_by_collector`` (``dict[str, list[dict]]``).
    All other runners save ``results`` (``list[dict]``) which gets wrapped
    under a synthetic ``"_flat"`` key so callers always get the same shape.
    """
    if "results_by_collector" in entry:
        return entry["results_by_collector"]
    flat = entry.get("results", [])
    if flat:
        kind = entry.get("kind", "_flat")
        return {kind: flat}
    return {}


def get_urls_from_results(entry: dict) -> list[str]:
    """Extract all unique URLs from a history entry, regardless of format."""
    urls: list[str] = []
    by_collector = get_results(entry)
    for rows in by_collector.values():
        for r in rows:
            u = r.get("url") or r.get("source_url") or ""
            if u:
                urls.append(u)
    return list(dict.fromkeys(urls))


class CaptureRunner:
    def __init__(self, urls: list[str], runtime_cfg: dict, output_dir: Path, kind: str = "screenshot", generate_pdf: bool = False, seo_fields: list[dict] | None = None, progress_callback: Callable[[int, int, str], None] | None = None):
        self.urls = urls
        self.runtime_cfg = runtime_cfg
        self.output_dir = output_dir
        self.kind = kind
        self.generate_pdf = generate_pdf
        self.seo_fields = seo_fields
        self.progress_callback = progress_callback
        self.results = []
        self.cancelled = False
        self._thread = None
        self.progress_total = len(urls)
        self.progress_done = 0
        self.status = "queued"

    def _report_progress(self) -> None:
        if self.progress_callback:
            self.progress_callback(self.progress_done, self.progress_total, self.status)

    def run(self):
        self.results = []
        kind = self.kind
        output_dir = self.output_dir
        urls = self.urls
        runtime_cfg = self.runtime_cfg
        photos_dir = output_dir / "photos"
        photos_dir.mkdir(parents=True, exist_ok=True)
        data_dir = output_dir / "data"
        data_dir.mkdir(parents=True, exist_ok=True)
        results = self.results
        self.status = "Running..."
        self._report_progress()

        with SB(uc=True, test=True, headless=False, window_size=f"{runtime_cfg['viewport']['width']},{runtime_cfg['viewport']['height']}") as sb:
            page = PageCapture(sb, runtime_cfg)
            for i, url in enumerate(urls):
                if self.cancelled:
                    break
                self.status = f"Processing {url}"
                slug = slugify(url)
                row = {"url": url, "status": "waiting"}
                try:
                    if not is_valid_url(url):
                        row["status"] = "invalid URL"
                        results.append(row)
                        continue
                    page.open(url)
                    page.scroll()
                    sb.sleep(runtime_cfg["timing"]["stabilization_ms"] / 1000)
                    page.hide_overlays()
                    if kind == "seo":
                        fields = self.seo_fields or get_standard_seo_fields()
                        raw = sb.cdp.evaluate(build_seo_js(fields))
                        row = {"url": url, "status": "ok", **parse_seo_payload(raw, fields)}
                    else:
                        png_path = photos_dir / f"{slug}.png"
                        page.capture_png(png_path)
                        pdf_path = None
                        if self.generate_pdf:
                            pdf_dir = output_dir / "pdf"
                            pdf_dir.mkdir(parents=True, exist_ok=True)
                            pdf_path = pdf_dir / f"{slug}.pdf"
                            png_to_pdf(png_path, pdf_path)
                        page_data = page.extract_data()
                        row = {
                            "url": url, "status": "ok",
                            "page_name": page_data.get("page_name", ""),
                            "h1": page_data.get("h1", ""),
                            "file": str(png_path),
                            "pdf": str(pdf_path) if pdf_path else "",
                        }
                except Exception as exc:
                    row = {"url": url, "status": f"error: {exc}"}
                results.append(row)
                self.progress_done += 1
                self._report_progress()
                time.sleep(random.uniform(
                    runtime_cfg["timing"]["inter_page_delay_min"],
                    runtime_cfg["timing"]["inter_page_delay_max"],
                ))

        self.status = "done"
        self._report_progress()

        # Compute internal_inlinks from outlink counts across all pages
        if kind == "seo":
            _compute_internal_inlinks(results)

        csv_path = data_dir / ("seo_results.csv" if kind == "seo" else "capture_results.csv")
        write_results_csv(results, csv_path)

        timestamp = datetime.now().isoformat()
        total = len(results)
        ok_count = sum(1 for r in results if r.get("status") == "ok")
        save_history({
            "timestamp": timestamp,
            "kind": kind,
            "total": total,
            "ok": ok_count,
            "fail": total - ok_count,
            "output_dir": str(output_dir),
            "results": results,
        })


class ExtractionRunner:
    def __init__(self, urls: list[str], rules: list[dict], runtime_cfg: dict, output_dir: Path, progress_callback: Callable[[int, int, str], None] | None = None):
        self.urls = urls
        self.rules = rules
        self.runtime_cfg = runtime_cfg
        self.output_dir = output_dir
        self.progress_callback = progress_callback
        self.results = []
        self.cancelled = False
        self._thread = None
        self.progress_total = len(urls)
        self.progress_done = 0
        self.status = "queued"

    def _report_progress(self) -> None:
        if self.progress_callback:
            self.progress_callback(self.progress_done, self.progress_total, self.status)

    def run(self):
        self.results = []
        output_dir = self.output_dir
        urls = self.urls
        rules = self.rules
        runtime_cfg = self.runtime_cfg
        data_dir = output_dir / "data"
        data_dir.mkdir(parents=True, exist_ok=True)
        results = self.results
        self.status = "Running..."
        self._report_progress()

        with SB(uc=True, test=True, headless=False, window_size=f"{runtime_cfg['viewport']['width']},{runtime_cfg['viewport']['height']}") as sb:
            page = PageCapture(sb, runtime_cfg)
            for i, url in enumerate(urls):
                if self.cancelled:
                    break
                self.status = f"Processing {url}"
                row = {"url": url, "status": "waiting"}
                try:
                    if not is_valid_url(url):
                        row["status"] = "invalid URL"
                        results.append(row)
                        continue
                    page.open(url)
                    page.scroll()
                    sb.sleep(runtime_cfg["timing"]["stabilization_ms"] / 1000)
                    page.hide_overlays()
                    data = extract_from_page(sb, rules)
                    row = {"url": url, "status": "ok", **data}
                except Exception as exc:
                    row = {"url": url, "status": f"error: {exc}"}
                results.append(row)
                self.progress_done += 1
                self._report_progress()
                time.sleep(random.uniform(
                    runtime_cfg["timing"]["inter_page_delay_min"],
                    runtime_cfg["timing"]["inter_page_delay_max"],
                ))

        self.status = "done"
        self._report_progress()

        csv_path = data_dir / "extraction_results.csv"
        write_results_csv(results, csv_path)

        timestamp = datetime.now().isoformat()
        total = len(results)
        ok_count = sum(1 for r in results if r.get("status") == "ok")
        save_history({
            "timestamp": timestamp,
            "kind": "extraction",
            "total": total,
            "ok": ok_count,
            "fail": total - ok_count,
            "output_dir": str(output_dir),
            "results": results,
            "extraction_rules": self.rules,
        })


class UnifiedRunner:
    """Run multiple collectors against the same URL list in one browser session.

    Each collector produces its own CSV; the history entry is one record tagged
    with the list of collectors actually run.
    """

    _thread: Optional[threading.Thread]

    def __init__(self, urls: list[str], collectors: list[dict], runtime_cfg: dict, output_dir: Path, seo_fields: list[dict] | None = None, generate_pdf: bool = False, progress_callback: Callable[[int, int, str], None] | None = None):
        self.urls = urls
        self.collectors = collectors
        self.runtime_cfg = runtime_cfg
        self.output_dir = output_dir
        self.seo_fields = seo_fields
        self.generate_pdf = generate_pdf
        self.progress_callback = progress_callback
        self.results = {"screenshot": [], "seo": [], "extraction": []}
        self.cancelled = False
        self._thread = None
        self.status = "queued"
        self.progress_total = len(urls) * max(len(collectors), 1)
        self.progress_done = 0

    def _bump_progress(self, n: int = 1) -> None:
        self.progress_done += n
        if self.progress_callback:
            self.progress_callback(self.progress_done, self.progress_total, self.status)

    def run(self):
        self.results = {"screenshot": [], "seo": [], "extraction": []}
        self.progress_done = 0
        urls = self.urls
        runtime_cfg = self.runtime_cfg
        output_dir = self.output_dir
        collectors = self.collectors

        photos_dir = output_dir / "photos"
        photos_dir.mkdir(parents=True, exist_ok=True)
        data_dir = output_dir / "data"
        data_dir.mkdir(parents=True, exist_ok=True)

        active = {c["name"] for c in collectors}
        run_screenshot = "screenshot" in active
        run_seo = "seo" in active
        run_extraction = "extraction" in active

        rules_by_collector: dict[str, list[dict]] = {}
        for c in collectors:
            if c["name"] == "extraction":
                rules_by_collector["extraction"] = c.get("rules") or []

        with SB(uc=True, test=True, headless=False, window_size=f"{runtime_cfg['viewport']['width']},{runtime_cfg['viewport']['height']}") as sb:
            page = PageCapture(sb, runtime_cfg)
            for i, url in enumerate(urls):
                if self.cancelled:
                    break
                slug = slugify(url)

                ss_row = None
                seo_row = None
                ext_row = None

                ok = False
                err = None
                if not is_valid_url(url):
                    err = "invalid URL"
                else:
                    try:
                        self.status = f"Opening {url}"
                        page.open(url)
                        page.scroll()
                        sb.sleep(runtime_cfg["timing"]["stabilization_ms"] / 1000)
                        page.hide_overlays()
                        ok = True
                    except Exception as exc:
                        err = str(exc)

                if run_screenshot:
                    if ok:
                        try:
                            png_path = photos_dir / f"{slug}.png"
                            page.capture_png(png_path)
                            page_data = page.extract_data()
                            ss_row = {
                                "url": url, "status": "ok",
                                "page_name": page_data.get("page_name", ""),
                                "h1": page_data.get("h1", ""),
                                "file": str(png_path),
                            }
                            if self.generate_pdf:
                                pdf_dir = output_dir / "pdf"
                                pdf_dir.mkdir(parents=True, exist_ok=True)
                                pdf_path = pdf_dir / f"{slug}.pdf"
                                png_to_pdf(png_path, pdf_path)
                                ss_row["pdf"] = str(pdf_path)
                        except Exception as exc:
                            ss_row = {"url": url, "status": f"error: {exc}"}
                    else:
                        ss_row = {"url": url, "status": f"error: {err}"}
                    self.results["screenshot"].append(ss_row)
                    self._bump_progress()

                if run_seo and not self.cancelled:
                    if ok:
                        try:
                            fields = self.seo_fields or get_standard_seo_fields()
                            raw = sb.cdp.evaluate(build_seo_js(fields))
                            seo_row = {"url": url, "status": "ok", **parse_seo_fields(raw, fields)}
                        except Exception as exc:
                            seo_row = {"url": url, "status": f"error: {exc}"}
                    else:
                        seo_row = {"url": url, "status": f"error: {err}"}
                    self.results["seo"].append(seo_row)
                    self._bump_progress()

                if run_extraction and not self.cancelled:
                    rules = rules_by_collector.get("extraction", [])
                    if ok and rules:
                        try:
                            data = extract_from_page(sb, rules)
                            ext_row = {"url": url, "status": "ok", **data}
                        except Exception as exc:
                            ext_row = {"url": url, "status": f"error: {exc}"}
                    elif not ok:
                        ext_row = {"url": url, "status": f"error: {err}"}
                    else:
                        ext_row = {"url": url, "status": "no rules defined"}
                    self.results["extraction"].append(ext_row)
                    self._bump_progress()

                time.sleep(random.uniform(
                    runtime_cfg["timing"]["inter_page_delay_min"],
                    runtime_cfg["timing"]["inter_page_delay_max"],
                ))

        # Compute internal_inlinks from outlink counts across all pages
        if self.results.get("seo"):
            _compute_internal_inlinks(self.results["seo"])

        for kind, rows in self.results.items():
            if not rows:
                continue
            csv_path = data_dir / f"{kind}_results.csv"
            write_results_csv(rows, csv_path)

        total = sum(len(r) for r in self.results.values())
        ok_count = sum(
            1 for rows in self.results.values() for r in rows if r.get("status") == "ok"
        )
        extraction_rules = []
        for c in collectors:
            if c["name"] == "extraction":
                extraction_rules = c.get("rules") or []
        save_history({
            "timestamp": datetime.now().isoformat(),
            "kind": "unified",
            "collectors": [c["name"] for c in collectors],
            "total": total,
            "ok": ok_count,
            "fail": total - ok_count,
            "output_dir": str(output_dir),
            "results_by_collector": self.results,
            "extraction_rules": extraction_rules,
            "fast_mode": False,
        })
        self.status = "done"


# ── Fast crawl (curl_cffi + threads) ────────────────────────────────────────────


def _fetch_and_extract(
    url: str,
    cookies: dict[str, str],
    user_agent: str,
    timeout: float = 30.0,
) -> dict:
    """Fetch one URL with curl_cffi (Chrome TLS impersonation) and extract SEO data."""
    from curl_cffi import requests as _curl
    from lxml import html as _html

    headers = {"User-Agent": user_agent} if user_agent else {}
    try:
        resp = _curl.get(url, cookies=cookies, headers=headers, timeout=timeout, impersonate="chrome")
    except Exception as exc:
        return {"url": url, "status": f"error: {exc}"}

    status_code = resp.status_code
    host = url.split("//")[-1].split("/")[0].split(":")[0]
    # Non-2xx responses are almost always bot-block pages (403/503/etc.)
    if status_code != 200:
        return {
            "url": url,
            "status": f"http_{status_code}",
            "status_code": status_code,
        }
    data: dict = {"url": url, "status": "ok", "status_code": status_code}

    try:
        tree = _html.fromstring(resp.text)
    except Exception:
        tree = None

    if tree is None:
        return data

    def _text(expr: str) -> str:
        return " ".join(t.strip() for t in tree.xpath(expr) if isinstance(t, str) and t.strip())

    def _attr(expr: str) -> str:
        vals = tree.xpath(expr)
        return vals[0].strip() if vals else ""

    title = _text("//title/text()")
    data["title"] = title
    data["title_len"] = len(title)

    meta_desc = _attr("//meta[@name='description']/@content")
    data["meta_description"] = meta_desc
    data["meta_desc_len"] = len(meta_desc)

    data["canonical"] = _attr("//link[@rel='canonical']/@href")
    data["robots_meta"] = _attr("//meta[@name='robots']/@content")

    data["h1"] = _text("//h1//text()")
    data["h2s"] = " | ".join(t for t in _text("//h2//text()").split(" | ")[:15])[:500]
    data["h3s"] = " | ".join(t for t in _text("//h3//text()").split(" | ")[:15])[:500]

    data["og_title"] = _attr("//meta[@property='og:title']/@content")
    data["og_description"] = _attr("//meta[@property='og:description']/@content")
    data["og_image"] = _attr("//meta[@property='og:image']/@content")

    schema_types = []
    for script_text in tree.xpath('//script[@type="application/ld+json"]/text()'):
        try:
            d = json.loads(script_text)
            t = d.get("@type", "")
            if isinstance(t, list):
                schema_types.extend(t)
            elif t:
                schema_types.append(t)
        except (json.JSONDecodeError, AttributeError):
            pass
    data["schema_types"] = " | ".join(schema_types)

    body_text = " ".join(tree.xpath("//body//text()"))
    data["word_count"] = len(body_text.split())

    internal = 0
    external = 0
    for href in tree.xpath("//a/@href"):
        href = href.strip()
        if href.startswith("#") or href.startswith("mailto:") or href.startswith("tel:"):
            continue
        if href.startswith("//"):
            link_host = href[2:].split("/")[0].split(":")[0]
        elif href.startswith("http"):
            link_host = href.split("//")[-1].split("/")[0].split(":")[0]
        else:
            link_host = host
        if link_host == host:
            internal += 1
        else:
            external += 1
    data["internal_links"] = internal
    data["external_links"] = external

    imgs = tree.xpath("//img")
    missing = sum(1 for img in imgs if not (img.attrib.get("alt", "") or "").strip())
    data["images_missing_alt"] = missing

    # Soft-block detection: sites sometimes return 200 with a challenge/block page.
    # (Hard blocks via non-200 are already caught above by status_code != 200.)
    _BLOCK_MARKERS = (
        "just a moment", "attention required", "checking your browser",
        "dealer website", "access denied", "are you a robot", "enable javascript",
    )
    if data.get("status") == "ok":
        title_low = (data.get("title") or "").lower()
        if any(m in title_low for m in _BLOCK_MARKERS):
            data["status"] = "blocked"

    return data


class FastRunnerLegacy:
    """Fast crawl using curl_cffi + ThreadPoolExecutor with cookies from SeleniumBase.

    Flow:
    1. Open first URL in SeleniumBase, solve Turnstile.
    2. Export cookies + user-agent.
    3. Close browser.
    4. Crawl all URLs concurrently with curl_cffi (8 threads).
    5. Collect results into ``self.results["seo"]``.
    """

    _thread: Optional[threading.Thread]

    def __init__(
        self,
        urls: list[str],
        runtime_cfg: dict,
        output_dir: Path,
        seo_fields: list[dict] | None = None,
        progress_callback: Callable[[int, int, str], None] | None = None,
    ):
        self.urls = urls
        self.runtime_cfg = runtime_cfg
        self.output_dir = output_dir
        self.seo_fields = seo_fields
        self.progress_callback = progress_callback
        self.results: dict[str, list[dict]] = {"seo": []}
        self.cancelled = False
        self._thread = None
        self.status = "queued"
        self.progress_total = len(urls)
        self.progress_done = 0

    def _report_progress(self) -> None:
        if self.progress_callback:
            self.progress_callback(self.progress_done, self.progress_total, self.status)

    def _refresh_session(self, seed_url: str) -> dict:
        """Open a browser, solve Turnstile, and return a fresh session dict.

        Used to obtain a new clearance cookie when a crawl batch gets blocked.
        """
        viewport = self.runtime_cfg["viewport"]
        try:
            with SB(
                uc=True, test=True, headless=False,
                window_size=f"{viewport['width']},{viewport['height']}",
            ) as sb:
                page = PageCapture(sb, self.runtime_cfg)
                self.status = f"Solving Turnstile on {seed_url}"
                page.open(seed_url)
                page.scroll()
                sb.sleep(self.runtime_cfg["timing"]["stabilization_ms"] / 1000)
                page.hide_overlays()
                return page.extract_session()
        except Exception:
            return {}

    def _crawl_batch(self, urls: list[str], cookies_dict: dict, user_agent: str) -> list[dict]:
        if not urls:
            return []
        items: list[dict] = []
        max_workers = min(8, len(urls))

        def _crawl_one(url: str) -> dict:
            try:
                return _fetch_and_extract(url, cookies_dict, user_agent)
            except Exception as exc:
                return {"url": url, "status": f"error: {exc}"}

        with ThreadPoolExecutor(max_workers=max_workers) as pool:
            futures = {pool.submit(_crawl_one, url): url for url in urls}
            for future in as_completed(futures):
                if self.cancelled:
                    break
                items.append(future.result())
        return items

    def run(self):
        self.results = {"seo": []}
        self.progress_done = 0
        self.status = "Starting browser..."
        self._report_progress()

        output_dir = self.output_dir
        data_dir = output_dir / "data"
        data_dir.mkdir(parents=True, exist_ok=True)

        # Step 1: open browser, solve Turnstile on first URL, export session
        session = self._refresh_session(self.urls[0])
        cookies_dict = {c["name"]: c["value"] for c in session.get("cookies", [])}
        user_agent = session.get("user_agent", "")

        # Step 2: crawl all URLs concurrently with curl_cffi
        self.status = f"Crawling {len(self.urls)} URLs..."
        self._report_progress()
        items = self._crawl_batch(self.urls, cookies_dict, user_agent)
        self.progress_done = min(len(items), self.progress_total)
        self._report_progress()

        # Step 3: retry anything that was blocked with a fresh Turnstile session
        max_retries = 3
        for attempt in range(max_retries):
            blocked = [
                it for it in items
                if not (it.get("status") == "ok" and it.get("status_code") == 200)
            ]
            if not blocked or self.cancelled:
                break
            self.status = f"Retrying {len(blocked)} blocked URLs (attempt {attempt + 1}/{max_retries})..."
            self._report_progress()
            fresh = self._refresh_session(blocked[0]["url"])
            if not fresh.get("cookies"):
                break
            cookies_dict = {c["name"]: c["value"] for c in fresh.get("cookies", [])}
            user_agent = fresh.get("user_agent", "")
            retried = self._crawl_batch([it["url"] for it in blocked], cookies_dict, user_agent)
            by_url = {r["url"]: r for r in retried}
            for it in items:
                u = it.get("url")
                if u in by_url:
                    it.clear()
                    it.update(by_url[u])
            self.progress_done = min(len(items), self.progress_total)
            self._report_progress()

        self.results["seo"] = items
        self.progress_done = self.progress_total
        self.status = "done"
        self._report_progress()

        if items:
            _compute_internal_inlinks(items)

        csv_path = data_dir / "seo_results.csv"
        write_results_csv(items, csv_path)

        total = len(items)
        ok_count = sum(1 for r in items if r.get("status") == "ok")
        save_history({
            "timestamp": datetime.now().isoformat(),
            "kind": "fast_seo",
            "total": total,
            "ok": ok_count,
            "fail": total - ok_count,
            "output_dir": str(output_dir),
            "results": items,
            "fast_mode": True,
            "collectors": ["seo"],
        })


# ── Crawl4AI Runner ────────────────────────────────────────────────────────────


class Crawl4AIRunner:
    """Fast SEO crawl using Crawl4AI (async Playwright + structured output)."""

    _thread: Optional[threading.Thread]

    def __init__(
        self,
        urls: list[str],
        runtime_cfg: dict,
        output_dir: Path,
        seo_fields: list[dict] | None = None,
        progress_callback: Callable[[int, int, str], None] | None = None,
        crawl_config: dict | None = None,
    ):
        self.urls = urls
        self.runtime_cfg = runtime_cfg
        self.output_dir = output_dir
        self.seo_fields = seo_fields  # kept for compatibility; crawl4ai returns all fields
        self.progress_callback = progress_callback
        self.results: dict[str, list[dict]] = {"seo": []}
        self.cancelled = False
        self._thread = None
        self.status = "queued"
        self.progress_total = len(urls)
        self.progress_done = 0
        self.crawl_config = crawl_config or {}

    def _crawl4ai_config(self) -> dict:
        """Build Crawl4AI configuration from runtime config and crawl4ai.yaml."""
        c4ai = self.runtime_cfg.get("crawl4ai", {})
        cc = self.crawl_config
        return {
            "rate_limit": (c4ai.get("rate_limit_rps", 10), c4ai.get("rate_limit_burst", 1)),
            "timeout": c4ai.get("timeout", 30),
            "wait_until": c4ai.get("wait_until", "domcontentloaded"),
            "wait_for_timeout": c4ai.get("wait_for_timeout", 15000),
            "headless": c4ai.get("headless", True),
            "viewport_width": c4ai.get("viewport_width", self.runtime_cfg["viewport"]["width"]),
            "viewport_height": c4ai.get("viewport_height", self.runtime_cfg["viewport"]["height"]),
            "user_agent": c4ai.get("user_agent", "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"),
            "simulate_user": c4ai.get("simulate_user", True),
            "magic": c4ai.get("magic", True),
            "override_navigator": c4ai.get("override_navigator", True),
            "user_agent_mode": c4ai.get("user_agent_mode", "random"),
            "session_id": c4ai.get("session_id", "default"),
            "mean_delay": c4ai.get("mean_delay", 1.0),
            "max_range": c4ai.get("max_range", 2.0),
            "navigation_timeout": c4ai.get("navigation_timeout", 60000),
            # Step 1: Core crawl configuration
            "max_depth": cc.get("max_depth", c4ai.get("max_depth", 0)),
            "max_pages": cc.get("max_pages", c4ai.get("max_pages", 1000)),
            "include_patterns": cc.get("include_patterns", c4ai.get("include_patterns", [])),
            "exclude_patterns": cc.get("exclude_patterns", c4ai.get("exclude_patterns", [])),
            "strip_query_params": cc.get("strip_query_params", c4ai.get("strip_query_params", False)),
            "respect_robots_txt": cc.get("respect_robots_txt", c4ai.get("respect_robots_txt", False)),
            "allowed_domains": cc.get("allowed_domains", c4ai.get("allowed_domains", [])),
            "blocked_domains": cc.get("blocked_domains", c4ai.get("blocked_domains", [])),
        }

    def _transform_result(self, result) -> dict:
        """Transform Crawl4AI CrawlResult to existing SEO dict format."""
        if result is None:
            return {
                "url": "", "status": "error: no result", "status_code": 0,
                "title": "", "title_len": 0, "meta_description": "", "meta_desc_len": 0,
                "canonical": "", "robots_meta": "", "h1": "", "h2s": "", "h3s": "",
                "og_title": "", "og_description": "", "og_image": "", "og_type": "",
                "og_url": "", "og_site_name": "", "og_locale": "",
                "twitter_card": "", "twitter_title": "", "twitter_description": "",
                "twitter_image": "", "twitter_site": "", "schema_types": "",
                "word_count": 0, "internal_links": 0, "external_links": 0,
                "images_missing_alt": 0, "images_total": 0, "images_no_lazy": 0,
                "iframe_count": 0, "form_count": 0, "external_nofollow": 0,
                "html_lang": "", "meta_viewport": "", "meta_charset": "",
                "hreflang": "", "jsonld_full": "",
                "depth": 0, "redirect_chain": "", "page_size": 0, "response_time": 0,
                "redirected_url": "", "redirected_status_code": None,
                "canonical_chain": "", "internal_inlinks": 0,
                "internal_links_detail": [], "external_links_detail": [],
                "content_type": "", "images_detail": [], "boilerplate_ratio": 0.0,
            }

        md = result.metadata or {}
        schema_types = md.get("schema", []) or []

        url = getattr(result, "url", "") or ""
        success = getattr(result, "success", False)
        error_msg = getattr(result, "error_message", None) or "unknown"
        status_code = getattr(result, "status_code", 0) or 0
        depth = getattr(result, "depth", 0) or md.get("depth", 0)
        page_size = md.get("page_size", 0) or 0
        response_time = md.get("response_time", 0) or 0

        # ── Step 3: Link and Redirect Analysis ──

        redirected_url = getattr(result, "redirected_url", None) or ""
        redirected_status_code = getattr(result, "redirected_status_code", None)

        # Build redirect chain from redirected_url info
        redirect_chain = md.get("redirect_chain", []) or []
        if redirected_url and redirected_url != url:
            chain_entry = f"{redirected_url} ({redirected_status_code or '?'})"
            if chain_entry not in redirect_chain:
                redirect_chain.append(chain_entry)

        # Store individual link details from result.links
        raw_links = getattr(result, "links", None) or {}
        internal_links_detail: list[dict] = []
        external_links_detail: list[dict] = []
        for entry in raw_links.get("internal", []):
            if isinstance(entry, dict):
                internal_links_detail.append({
                    "href": entry.get("href", ""),
                    "text": entry.get("text", "") or entry.get("content", ""),
                    "rel": entry.get("rel", []),
                })
        for entry in raw_links.get("external", []):
            if isinstance(entry, dict):
                external_links_detail.append({
                    "href": entry.get("href", ""),
                    "text": entry.get("text", "") or entry.get("content", ""),
                    "rel": entry.get("rel", []),
                })

        # Canonical chain: track resolved chain for canonical URLs
        canonical_url = md.get("canonical", "")
        canonical_chain = []
        if canonical_url and canonical_url != url:
            canonical_chain.append(canonical_url)

        # ── Step 4: On-Page SEO Auditing ──

        # Extract image details from result.media
        raw_media = getattr(result, "media", None) or {}
        images_detail: list[dict] = []
        for img in raw_media.get("images", []):
            if isinstance(img, dict):
                images_detail.append({
                    "src": img.get("src", "") or img.get("url", ""),
                    "alt": img.get("alt", "") or "",
                    "width": img.get("width", 0) or 0,
                    "height": img.get("height", 0) or 0,
                    "filesize": img.get("filesize", 0) or img.get("size", 0),
                    "format": img.get("type", "") or img.get("format", ""),
                })

        # Content type from response headers
        resp_headers = getattr(result, "response_headers", None) or {}
        content_type = (resp_headers.get("content-type", resp_headers.get("Content-Type", "")) or "").split(";")[0].strip()

        return {
            "url": url,
            "status": "ok" if success else f"error: {error_msg}",
            "status_code": status_code,
            "title": md.get("title", ""),
            "title_len": len(md.get("title", "")),
            "meta_description": md.get("description") or "",
            "meta_desc_len": len(md.get("description") or ""),
            "canonical": canonical_url,
            "robots_meta": md.get("robots", ""),
            "h1": md.get("h1", ""),
            "h2s": " | ".join(md.get("h2", [])) if isinstance(md.get("h2"), list) else str(md.get("h2", "")),
            "h3s": " | ".join(md.get("h3", [])) if isinstance(md.get("h3"), list) else str(md.get("h3", "")),
            "og_title": md.get("og_title", ""),
            "og_description": md.get("og_description", ""),
            "og_image": md.get("og_image", ""),
            "og_type": md.get("og_type", ""),
            "og_url": md.get("og_url", ""),
            "og_site_name": md.get("og_site_name", ""),
            "og_locale": md.get("og_locale", ""),
            "twitter_card": md.get("twitter_card", ""),
            "twitter_description": md.get("twitter_description", ""),
            "twitter_image": md.get("twitter_image", ""),
            "twitter_site": md.get("twitter_site", ""),
            "schema_types": " | ".join(schema_types) if isinstance(schema_types, list) else str(schema_types),
            "word_count": md.get("word_count", 0),
            "internal_links": len(internal_links_detail),
            "external_links": len(external_links_detail),
            "internal_links_detail": internal_links_detail,
            "external_links_detail": external_links_detail,
            "images_missing_alt": md.get("images_missing_alt", 0),
            "images_total": md.get("images_total", 0),
            "images_no_lazy": md.get("images_no_lazy", 0),
            "iframe_count": md.get("iframe_count", 0),
            "form_count": md.get("form_count", 0),
            "external_nofollow": md.get("external_nofollow", 0),
            "html_lang": md.get("html_lang", ""),
            "meta_viewport": md.get("meta_viewport", ""),
            "meta_charset": md.get("meta_charset", ""),
            "hreflang": md.get("hreflang", ""),
            "jsonld_full": md.get("jsonld_full", ""),
            # Step 1: Core crawl fields
            "depth": depth,
            "redirect_chain": " -> ".join(redirect_chain) if redirect_chain else "",
            "page_size": page_size,
            "response_time": response_time,
            # Step 3: Link and redirect details
            "redirected_url": redirected_url,
            "redirected_status_code": redirected_status_code,
            "canonical_chain": " -> ".join(canonical_chain) if canonical_chain else "",
            "internal_inlinks": 0,
            # Step 4: On-Page SEO details
            "content_type": content_type,
            "images_detail": images_detail,
            "boilerplate_ratio": 0.0,
        }

    def _report_progress(self) -> None:
        if self.progress_callback:
            self.progress_callback(self.progress_done, self.progress_total, self.status)

    def run(self):
        import asyncio

        from crawl4ai import AsyncWebCrawler, CacheMode, CrawlerRunConfig
        from crawl4ai.deep_crawling import BFSDeepCrawlStrategy
        from crawl4ai.deep_crawling.filters import (
            DomainFilter,
            FilterChain,
            URLPatternFilter,
        )

        self.results = {"seo": []}
        self.progress_done = 0
        self.status = "Starting Crawl4AI..."
        self._report_progress()

        output_dir = self.output_dir
        data_dir = output_dir / "data"
        data_dir.mkdir(parents=True, exist_ok=True)

        cfg = self._crawl4ai_config()

        run_cfg = CrawlerRunConfig(
            cache_mode=CacheMode.BYPASS,
            wait_until=cfg["wait_until"],
            wait_for_timeout=cfg["wait_for_timeout"],
            page_timeout=cfg["timeout"] * 1000,
            session_id=cfg.get("session_id") or "persistent_session",
            mean_delay=cfg.get("mean_delay", 1.0),
            max_range=cfg.get("max_range", 2.0),
            override_navigator=cfg.get("override_navigator", True),
            user_agent_mode=cfg.get("user_agent_mode", "random"),
            remove_overlay_elements=cfg.get("remove_overlay_elements", True),
            remove_consent_popups=cfg.get("remove_consent_popups", True),
            check_robots_txt=cfg.get("respect_robots_txt", False),
        )

        max_depth = cfg.get("max_depth", 0)
        max_pages = cfg.get("max_pages", 1000)
        include_patterns: list[str] = cfg.get("include_patterns", [])
        exclude_patterns: list[str] = cfg.get("exclude_patterns", [])
        allowed_domains: list[str] = cfg.get("allowed_domains", [])
        blocked_domains: list[str] = cfg.get("blocked_domains", [])
        strip_query_params = cfg.get("strip_query_params", False)

        async def _run_all():
            # ---- Batch crawl (max_depth=0): crawl exact URLs list ----

            async def _batch_crawl(crawler):
                async def _crawl(urls: list[str]):
                    try:
                        results = await crawler.arun_many(
                            urls,
                            config=run_cfg,
                            max_concurrent=cfg["rate_limit"][0] if isinstance(cfg["rate_limit"], tuple) else 8,
                        )
                    except Exception as e:
                        self.status = f"Navigation error: {e}, will retry..."
                        self._report_progress()
                        items = []
                        for url in urls:
                            if self.cancelled:
                                break
                            items.append({
                                "url": url,
                                "status": f"error: {e}",
                                "status_code": 0,
                            })
                        return items

                    if hasattr(results, "__aiter__"):
                        results_list = []
                        async for r in results:
                            results_list.append(r)
                        results = results_list

                    results = list(results)

                    items = []
                    for i, result in enumerate(results):
                        if self.cancelled:
                            break
                        self.progress_done = i + 1
                        self.status = f"Processing {result.url}"
                        self._report_progress()
                        row = self._transform_result(result)
                        items.append(row)

                    return items

                items = await _crawl(self.urls)

                max_retries = 3
                for attempt in range(max_retries):
                    blocked = [
                        it for it in items
                        if not (it.get("status") == "ok" and it.get("status_code") == 200 and it.get("title", "") and it.get("word_count", 0) > 0)
                    ]
                    if not blocked or self.cancelled:
                        break
                    self.status = f"Retrying {len(blocked)} failed URLs (attempt {attempt + 1}/{max_retries})..."
                    self._report_progress()
                    retry_urls = [it["url"] for it in blocked]

                    retried = await _crawl(retry_urls)

                    by_url = {r["url"]: r for r in retried}
                    for it in items:
                        u = it.get("url")
                        if u in by_url:
                            it.clear()
                            it.update(by_url[u])
                    self.progress_done = min(len(items), self.progress_total)
                    self._report_progress()

                return items

            # ---- Deep crawl (max_depth>0): BFS link following ----

            async def _deep_crawl(crawler):
                filters: list = []

                if include_patterns or exclude_patterns:
                    url_filter = URLPatternFilter(
                        patterns=[*include_patterns, *[f"!{p}" for p in exclude_patterns]],
                        use_glob=False,
                    )
                    filters.append(url_filter)

                if allowed_domains:
                    domain_filter = DomainFilter(allowed_domains=allowed_domains)
                    filters.append(domain_filter)
                if blocked_domains:
                    domain_filter = DomainFilter(blocked_domains=blocked_domains)
                    filters.append(domain_filter)

                filter_chain = FilterChain(filters=filters) if filters else None

                if not allowed_domains and not blocked_domains:
                    from urllib.parse import urlparse
                    seed_domains: list[str] = [h for h in {urlparse(u).hostname for u in self.urls} if h is not None]
                    if seed_domains:
                        domain_filter = DomainFilter(allowed_domains=seed_domains)
                        filter_chain = FilterChain(filters=[domain_filter])

                def _should_cancel():
                    return self.cancelled

                strategy = BFSDeepCrawlStrategy(
                    max_depth=max_depth,
                    filter_chain=filter_chain or FilterChain(filters=[]),
                    max_pages=max_pages,
                    should_cancel=_should_cancel,
                )

                crawl_cfg = CrawlerRunConfig(
                    cache_mode=CacheMode.BYPASS,
                    wait_until=cfg["wait_until"],
                    wait_for_timeout=cfg["wait_for_timeout"],
                    page_timeout=cfg["timeout"] * 1000,
                    session_id=cfg.get("session_id") or "persistent_session",
                    mean_delay=cfg.get("mean_delay", 1.0),
                    max_range=cfg.get("max_range", 2.0),
                    override_navigator=cfg.get("override_navigator", True),
                    user_agent_mode=cfg.get("user_agent_mode", "random"),
                    remove_overlay_elements=cfg.get("remove_overlay_elements", True),
                    remove_consent_popups=cfg.get("remove_consent_popups", True),
                    check_robots_txt=cfg.get("respect_robots_txt", False),
                    deep_crawl_strategy=strategy,
                )

                seen_urls: set[str] = set()
                all_items: list[dict] = []

                def _normalize_url(url: str) -> str:
                    if strip_query_params:
                        from urllib.parse import urlparse, urlunparse
                        parsed = urlparse(url)
                        return urlunparse(parsed._replace(query=""))
                    return url

                async def _crawl_seed(url: str) -> None:
                    container = await crawler.arun(url, config=crawl_cfg)
                    results_list: list = []
                    if hasattr(container, "__aiter__"):
                        async for r in container:
                            results_list.append(r)
                    elif hasattr(container, "__iter__"):
                        results_list = list(container)
                    else:
                        results_list = [container]

                    for result in results_list:
                        if self.cancelled:
                            break
                        u = getattr(result, "url", "")
                        normalized = _normalize_url(u)
                        if normalized and normalized not in seen_urls:
                            seen_urls.add(normalized)
                            row = self._transform_result(result)
                            row["depth"] = max_depth
                            all_items.append(row)
                            self.progress_done = len(all_items)
                            self.status = f"Crawled {u}"
                            self._report_progress()

                for seed_url in self.urls:
                    if self.cancelled:
                        break
                    self.status = f"Deep crawling from {seed_url}..."
                    self._report_progress()
                    await _crawl_seed(seed_url)

                self.progress_total = max(len(all_items), self.progress_total)
                return all_items

            # ---- Crawler execution ----

            async with AsyncWebCrawler(
                headless=cfg["headless"],
                verbose=False,
                viewport_width=cfg["viewport_width"],
                viewport_height=cfg["viewport_height"],
                user_agent=cfg["user_agent"],
                rate_limit=cfg["rate_limit"],
                simulate_user=cfg.get("simulate_user", True),
                magic=cfg.get("magic", True),
                override_navigator=cfg.get("override_navigator", True),
                user_agent_mode=cfg.get("user_agent_mode", "random"),
                remove_overlay_elements=cfg.get("remove_overlay_elements", True),
                remove_consent_popups=cfg.get("remove_consent_popups", True),
                proxy_config=cfg.get("proxy"),
            ) as crawler:
                if max_depth > 0:
                    return await _deep_crawl(crawler)
                else:
                    return await _batch_crawl(crawler)

        # Run async crawl in background thread
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            items = loop.run_until_complete(_run_all())
        finally:
            loop.close()

        self.results["seo"] = items
        self.progress_done = self.progress_total
        self.status = "done"
        self._report_progress()

        if items:
            _compute_internal_inlinks(items)
            self.status = "Analyzing broken links..."
            self._report_progress()
            broken = _analyze_broken_links(items)
            by_source: dict[str, list[dict]] = {}
            for entry in broken:
                by_source.setdefault(entry["source_url"], []).append(entry)
            for row in items:
                row["broken_links"] = by_source.get(row.get("url", ""), [])
            _resolve_canonical_chains(items)

            # Compute boilerplate ratio for each page
            for row in items:
                wc = row.get("word_count", 0) or 0
                ps = row.get("page_size", 0) or 0
                if ps > 0 and wc > 0:
                    estimated_content_bytes = wc * 6
                    row["boilerplate_ratio"] = round(max(0.0, 1.0 - estimated_content_bytes / ps), 2)

        # Step 5: Export results
        csv_path = data_dir / "seo_results.csv"
        write_results_csv(items, csv_path)
        excel_path = data_dir / "seo_results.xlsx"
        write_results_excel(items, excel_path)
        json_path = data_dir / "seo_results.json"
        write_results_json(items, json_path)

        total = len(items)
        ok_count = sum(1 for r in items if r.get("status") == "ok")
        stats = _compute_crawl_statistics(items) if items else {}
        issues_summary = {}
        if items:
            try:
                from analysis import analyze_results, summarize_issues
                issues_summary = summarize_issues(analyze_results(items))
            except Exception:
                issues_summary = {}
        save_history({
            "timestamp": datetime.now().isoformat(),
            "kind": "crawl4ai_seo",
            "total": total,
            "ok": ok_count,
            "fail": total - ok_count,
            "output_dir": str(output_dir),
            "results": items,
            "fast_mode": True,
            "collectors": ["seo"],
            "crawl_statistics": stats,
            "issues_summary": issues_summary,
            "crawl_config": self.crawl_config,
        })


# ── Blog Audit Runner ──────────────────────────────────────────────────────────


class BlogAuditRunner:
    """Compare blog posts between source and target sites using SeleniumBase CDP.

    For each URL pair:
      1. Load source page, extract blog data via CSS rules
      2. Load target page, extract blog data via same CSS rules
      3. Compare fields, score match quality
      4. Record issues (missing fields, unicode, image localization)

    Uses platform-specific extraction rulesets from rulesets/ directory.
    Defaults to generic_blog.json if no specific ruleset is provided.
    """

    _thread: Optional[threading.Thread]

    # Terms to exclude from category/tag extraction (site-wide noise)
    _TAXONOMY_EXCLUDE_TERMS = frozenset({
        "uncategorized", "blog", "all posts", "home", "about", "contact",
        "dealer", "dealership", "inventory", "service", "parts", "hours",
        "location", "directions", "finance", "specials", "reviews",
        "privacy", "sitemap", "careers", "testimonials", "team",
        "new inventory", "used inventory", "schedule service", "financing",
        "news", "events", "announcements", "press releases",
    })

    def __init__(
        self,
        url_pairs: list[tuple[str, str]],
        runtime_cfg: dict,
        output_dir: Path,
        progress_callback: Callable[[int, int, str], None] | None = None,
        ruleset_name: str = "generic_blog",
    ):
        self.url_pairs = url_pairs
        self.source_urls = [p[0] for p in url_pairs]
        self.target_urls = [p[1] for p in url_pairs]
        self.runtime_cfg = runtime_cfg
        self.output_dir = output_dir
        self.progress_callback = progress_callback
        self.results: dict[str, list[dict]] = {"audit": []}
        self.cancelled = False
        self._thread = None
        self.status = "queued"
        self.progress_total = len(url_pairs)
        self.progress_done = 0

        self._rules = []  # loaded at run()
        self._ruleset_name = ruleset_name
        self._slug_re = re.compile(r"/([^/]+)/?$")

    def _report_progress(self) -> None:
        if self.progress_callback:
            self.progress_callback(self.progress_done, self.progress_total, self.status)

    def _filter_taxonomy(self, items: list[str]) -> list[str]:
        """Filter categories/tags to exclude site-wide navigation terms and noise."""
        filtered = []
        for item in items:
            text = item.strip()
            if not text:
                continue
            lower = text.lower()
            if any(term in lower for term in self._TAXONOMY_EXCLUDE_TERMS):
                continue
            if "http" in lower or lower.startswith("/"):
                continue
            if len(text.split()) > 5:
                continue
            filtered.append(text)
        return filtered

    def _clean_content_html(self, html_str: str) -> str:
        """Post-process extracted HTML to remove unwanted elements and fix lazy images.

        Mimics the blog-tool's content cleanup pipeline.
        """
        if not html_str:
            return ""

        try:
            from bs4 import BeautifulSoup
        except ImportError:
            return html_str

        soup = BeautifulSoup(html_str, "html.parser")

        # Remove scripts, styles, noscript
        for tag in soup.find_all(["script", "style", "noscript"]):
            tag.decompose()

        # Remove breadcrumbs
        for sel in [".breadcrumbs", '.breadcrumb', 'nav[aria-label="Breadcrumb"]',
                     'nav[aria-label="breadcrumb"]']:
            for el in soup.select(sel):
                el.decompose()

        # Remove post navigation (prev/next links)
        for el in soup.select(".post-navigation, .nav-links, .posts-navigation"):
            el.decompose()

        # Remove duplicate title/date divs (dealer blog pattern)
        for el in soup.select(".titleDiv, .dateDiv, .content_title"):
            el.decompose()

        # Remove social sharing sections
        for el in soup.select(".sharingIcons, .social-share, .share-buttons, .post-share"):
            el.decompose()

        # Remove "Posted in" paragraphs
        for p in soup.find_all("p"):
            text = p.get_text(strip=True)
            if text.startswith("Posted in") or "Comments Off" in text or "comment is closed" in text.lower():
                p.decompose()

        # Remove "Connect with us" headings
        for heading in soup.find_all(["h2", "h3", "h4"]):
            if "connect with us" in heading.get_text().lower():
                heading.decompose()

        # Remove post metadata footer
        for el in soup.select(".postmetadata, .entry-meta, .post-meta"):
            el.decompose()

        # Fix lazy-loaded images: replace data-lazy-src, data-src with src
        for img in soup.find_all("img"):
            src = str(img.get("src") or "")
            if src and not src.startswith("data:"):
                continue
            for lazy_attr in ("data-lazy-src", "data-src", "data-original", "data-pin-media"):
                lazy_val = img.get(lazy_attr)
                if lazy_val:
                    img["src"] = lazy_val
                    break
            # Try srcset if still no good src
            cur_src = str(img.get("src") or "")
            if not cur_src or cur_src.startswith("data:"):
                srcset = img.get("data-lazy-srcset") or img.get("data-srcset") or img.get("srcset")
                if srcset:
                    candidates = [c.strip().split() for c in str(srcset).split(",")]
                    best = max(
                        (c for c in candidates if len(c) >= 2 and c[1].endswith("w")),
                        key=lambda c: int(c[1][:-1]),
                        default=None,
                    )
                    if best:
                        img["src"] = best[0]
            # Clean up lazy-loading attributes
            for attr in ("data-lazy-src", "data-src", "data-original", "data-pin-media",
                         "data-lazy-srcset", "data-srcset", "data-load-done", "data-ssr-src-done"):
                if attr in img.attrs:
                    del img[attr]

        result = str(soup)
        # Clean up double spaces / newlines
        import re
        result = re.sub(r"\n\s*\n", "\n", result)
        return result.strip()

    def _resolve_images_js(self) -> str:
        """Return a JS snippet that resolves lazy-loaded images in the DOM.

        Injects src from data-lazy-src/data-src/etc so our CSS selectors can
        find them before extraction runs.
        """
        return """
        (() => {
            document.querySelectorAll('img').forEach(img => {
                const src = img.getAttribute('src') || '';
                if (src && !src.startsWith('data:')) return;
                for (const attr of ['data-lazy-src', 'data-src', 'data-original', 'data-pin-media']) {
                    const val = img.getAttribute(attr);
                    if (val) { img.setAttribute('src', val); break; }
                }
                if (!img.getAttribute('src') || img.getAttribute('src').startsWith('data:')) {
                    const srcset = img.getAttribute('data-lazy-srcset') || img.getAttribute('data-srcset') || img.getAttribute('srcset');
                    if (srcset) {
                        let bestUrl = '', bestWidth = -1;
                        srcset.split(',').forEach(c => {
                            const parts = c.trim().split(' ');
                            if (parts.length >= 2 && parts[1].endsWith('w')) {
                                const w = parseInt(parts[1], 10);
                                if (w > bestWidth) { bestWidth = w; bestUrl = parts[0]; }
                            }
                        });
                        if (bestUrl) img.setAttribute('src', bestUrl);
                    }
                }
                for (const attr of ['data-lazy-src', 'data-src', 'data-original', 'data-pin-media',
                                     'data-lazy-srcset', 'data-srcset', 'data-load-done', 'data-ssr-src-done']) {
                    img.removeAttribute(attr);
                }
            });
        })()
        """

    def _extract_slug(self, url: str) -> str:
        m = self._slug_re.search(url.rstrip("/"))
        return m.group(1) if m else url

    def _extract_blog_data(self, page: PageCapture, url: str) -> dict:
        from extraction import extract_from_page

        page.open(url)
        page.scroll()
        page.sb.sleep(self.runtime_cfg["timing"]["stabilization_ms"] / 1000)
        page.hide_overlays()

        # Resolve lazy-loaded images in the DOM before extraction
        page.sb.cdp.evaluate(self._resolve_images_js())

        raw = extract_from_page(page.sb, self._rules)
        doc_title = (page.sb.cdp.evaluate("document.title") or "").strip()
        slug = self._extract_slug(url)

        # Post-process content HTML
        content_html = raw.get("content_html", "")
        cleaned_html = self._clean_content_html(content_html) if content_html else ""

        # Post-process categories/tags
        categories = self._filter_taxonomy(raw.get("categories", []))
        tags = self._filter_taxonomy(raw.get("tags", []))

        return {
            "url": url,
            "slug": slug,
            "doc_title": doc_title,
            "h1": raw.get("h1", ""),
            "meta_description": raw.get("meta_description", ""),
            "published_date": raw.get("published_date", ""),
            "author": raw.get("author", ""),
            "categories": categories,
            "tags": tags,
            "featured_image": raw.get("featured_image", ""),
            "content_html": cleaned_html,
            "content_text": raw.get("content_text", ""),
            "content_images": raw.get("content_images", []),
            "content_image_alts": raw.get("content_image_alts", []),
            "og_title": raw.get("og_title", ""),
            "og_image": raw.get("og_image", ""),
            "og_description": raw.get("og_description", ""),
        }

    def _compare_text(self, a: str, b: str) -> float:
        import difflib
        return difflib.SequenceMatcher(None, a.strip().lower(), b.strip().lower()).ratio()

    def _compare_sets(self, a: list, b: list) -> float:
        sa, sb = set(x.strip().lower() for x in a if x.strip()), set(x.strip().lower() for x in b if x.strip())
        if not sa and not sb:
            return 1.0
        union = sa | sb
        if not union:
            return 1.0
        intersection = sa & sb
        return len(intersection) / len(union)

    def _check_unicode(self, text: str) -> list[str]:
        issues = []
        if "\ufffd" in text:
            issues.append("Contains replacement character (�)")
        if "\u2013" in text or "\u2014" in text:
            pass  # en-dash / em-dash is fine
        for ch in text:
            if ord(ch) > 127 and ch not in "\u2013\u2014\u2018\u2019\u201c\u201d\u2026\u00a0\u00e9\u00e8\u00ea\u00eb\u00f1\u00fc\u00f6\u00e4\u00df\u00b0":
                if ord(ch) < 160:
                    issues.append(f"Unicode issue: char U+{ord(ch):04X} found")
        return issues

    def _check_image_localization(self, images: list[str], target_domain: str) -> list[str]:
        issues = []
        for img_url in images:
            if not img_url:
                continue
            try:
                from urllib.parse import urlparse
                parsed = urlparse(img_url)
                if parsed.netloc and parsed.netloc != target_domain:
                    if "dealeron" in parsed.netloc or "dealerdotinspire" in parsed.netloc:
                        issues.append(f"External image from {parsed.netloc}: {img_url[:80]}")
                    elif not img_url.startswith("data:"):
                        issues.append(f"External image from {parsed.netloc}: {img_url[:80]}")
            except Exception:
                pass
        return issues

    def _compare_entry(self, source: dict, target: dict) -> dict:
        fields = {}
        issues = []
        total_score = 0.0
        field_count = 0

        # Document title
        title_sim = self._compare_text(source.get("doc_title", ""), target.get("doc_title", ""))
        fields["title"] = {
            "source": source.get("doc_title", ""),
            "target": target.get("doc_title", ""),
            "score": round(title_sim * 100, 1),
            "match": title_sim > 0.85,
        }
        total_score += title_sim
        field_count += 1

        # H1
        h1_sim = self._compare_text(source.get("h1", ""), target.get("h1", ""))
        fields["h1"] = {
            "source": source.get("h1", ""),
            "target": target.get("h1", ""),
            "score": round(h1_sim * 100, 1),
            "match": h1_sim > 0.85,
        }
        total_score += h1_sim
        field_count += 1

        # Meta description
        desc_sim = self._compare_text(source.get("meta_description", ""), target.get("meta_description", ""))
        fields["meta_description"] = {
            "source": source.get("meta_description", ""),
            "target": target.get("meta_description", ""),
            "score": round(desc_sim * 100, 1),
            "match": desc_sim > 0.85 or (not source.get("meta_description") and not target.get("meta_description")),
        }
        total_score += desc_sim
        field_count += 1

        # Slug
        slug_sim = self._compare_text(source.get("slug", ""), target.get("slug", ""))
        fields["slug"] = {
            "source": source.get("slug", ""),
            "target": target.get("slug", ""),
            "score": round(slug_sim * 100, 1),
            "match": slug_sim > 0.95,
        }
        total_score += slug_sim
        field_count += 1

        # Date
        date_sim = self._compare_text(source.get("published_date", ""), target.get("published_date", ""))
        fields["published_date"] = {
            "source": source.get("published_date", ""),
            "target": target.get("published_date", ""),
            "score": round(date_sim * 100, 1),
            "match": date_sim > 0.9 or (not source.get("published_date") and not target.get("published_date")),
        }
        total_score += date_sim
        field_count += 1

        # Author
        author_sim = self._compare_text(source.get("author", ""), target.get("author", ""))
        fields["author"] = {
            "source": source.get("author", ""),
            "target": target.get("author", ""),
            "score": round(author_sim * 100, 1),
            "match": author_sim > 0.9 or (not source.get("author") and not target.get("author")),
        }
        total_score += author_sim
        field_count += 1

        # Categories
        cat_sim = self._compare_sets(source.get("categories", []), target.get("categories", []))
        fields["categories"] = {
            "source": source.get("categories", []),
            "target": target.get("categories", []),
            "score": round(cat_sim * 100, 1),
            "match": cat_sim > 0.8,
        }
        total_score += cat_sim
        field_count += 1

        # Tags
        tag_sim = self._compare_sets(source.get("tags", []), target.get("tags", []))
        fields["tags"] = {
            "source": source.get("tags", []),
            "target": target.get("tags", []),
            "score": round(tag_sim * 100, 1),
            "match": tag_sim > 0.8,
        }
        total_score += tag_sim
        field_count += 1

        # Featured image
        fi_sim = self._compare_text(source.get("featured_image", ""), target.get("featured_image", ""))
        fields["featured_image"] = {
            "source": source.get("featured_image", ""),
            "target": target.get("featured_image", ""),
            "score": round(fi_sim * 100, 1),
            "match": fi_sim > 0.9 or (not source.get("featured_image") and not target.get("featured_image")),
        }
        total_score += fi_sim
        field_count += 1

        # Content (using text similarity)
        src_text = source.get("content_text", "")
        tgt_text = target.get("content_text", "")
        content_sim = self._compare_text(src_text, tgt_text)
        src_html = source.get("content_html", "")
        tgt_html = target.get("content_html", "")

        # Image count comparison
        src_img_count = len(source.get("content_images", []))
        tgt_img_count = len(target.get("content_images", []))
        img_count_match = src_img_count == tgt_img_count

        fields["content"] = {
            "source_len": len(src_text),
            "target_len": len(tgt_text),
            "source_html_len": len(src_html),
            "target_html_len": len(tgt_html),
            "source_img_count": src_img_count,
            "target_img_count": tgt_img_count,
            "score": round(content_sim * 100, 1),
            "match": content_sim > 0.8,
        }
        total_score += content_sim
        field_count += 1

        if not img_count_match:
            issues.append({
                "field": "content_images",
                "type": "image_count_mismatch",
                "severity": "high",
                "message": f"Image count differs: source={src_img_count}, target={tgt_img_count}",
            })

        # Unicode checks
        for label, text_val in [
            ("title", source.get("doc_title", "")),
            ("title_target", target.get("doc_title", "")),
            ("h1", source.get("h1", "")),
            ("h1_target", target.get("h1", "")),
            ("content_source", src_text),
            ("content_target", tgt_text),
        ]:
            for issue in self._check_unicode(text_val):
                issues.append({
                    "field": label,
                    "type": "unicode",
                    "severity": "critical",
                    "message": issue,
                })

        # Image localization
        tgt_domain = ""
        try:
            from urllib.parse import urlparse
            tgt_domain = urlparse(target.get("url", "")).netloc
        except Exception:
            pass

        for img_url in target.get("content_images", []):
            for issue in self._check_image_localization([img_url], tgt_domain):
                issues.append({
                    "field": "content_images",
                    "type": "unlocalized_image",
                    "severity": "high",
                    "message": issue,
                })

        # Content empty check
        if src_text and not tgt_text:
            issues.append({
                "field": "content",
                "type": "missing",
                "severity": "critical",
                "message": "Target content is empty but source has content",
            })
        elif not src_text and tgt_text:
            issues.append({
                "field": "content",
                "type": "source_missing",
                "severity": "high",
                "message": "Source content is empty but target has content",
            })

        overall_score = round((total_score / max(field_count, 1)) * 100, 1)

        return {
            "source_url": source.get("url", ""),
            "target_url": target.get("url", ""),
            "overall_score": overall_score,
            "fields": fields,
            "issues": issues,
            "status": "ok" if overall_score >= 80 else "issues_found",
        }

    def run(self) -> None:
        self.results = {"audit": []}
        self.progress_done = 0
        self.status = "Starting browser..."
        self._report_progress()

        viewport = self.runtime_cfg["viewport"]
        from extraction import list_rulesets, load_ruleset
        available = list_rulesets()
        if self._ruleset_name not in available:
            self._ruleset_name = "generic_blog" if "generic_blog" in available else (available[0] if available else "")
        self._rules = load_ruleset(self._ruleset_name) or []

        if not self._rules:
            self.status = f"Error: No extraction rules found (tried: {self._ruleset_name})"
            self._report_progress()
            return

        self.status = f"Using ruleset: {self._ruleset_name}"

        self.status = f"Processing {len(self.url_pairs)} blog post pairs..."
        self._report_progress()

        audit_items: list[dict] = []

        try:
            with SB(
                uc=True, test=True, headless=False,
                window_size=f"{viewport['width']},{viewport['height']}",
            ) as sb:
                page = PageCapture(sb, self.runtime_cfg)

                for idx, (src_url, tgt_url) in enumerate(self.url_pairs):
                    if self.cancelled:
                        break

                    self.status = f"[{idx + 1}/{len(self.url_pairs)}] Source: {src_url}"
                    self._report_progress()

                    source_data = {}
                    target_data = {}

                    try:
                        source_data = self._extract_blog_data(page, src_url)
                    except Exception as exc:
                        source_data = {
                            "url": src_url,
                            "slug": self._extract_slug(src_url),
                            "status": f"error: {exc}",
                        }

                    self.status = f"[{idx + 1}/{len(self.url_pairs)}] Target: {tgt_url}"
                    self._report_progress()

                    try:
                        target_data = self._extract_blog_data(page, tgt_url)
                    except Exception as exc:
                        target_data = {
                            "url": tgt_url,
                            "slug": self._extract_slug(tgt_url),
                            "status": f"error: {exc}",
                        }

                    comparison = self._compare_entry(source_data, target_data)
                    audit_items.append(comparison)
                    self.progress_done = idx + 1
                    self._report_progress()

        except Exception as exc:
            self.status = f"Browser error: {exc}"
            self._report_progress()

        self.results["audit"] = audit_items
        self.progress_done = self.progress_total
        self.status = "done"
        self._report_progress()

        data_dir = self.output_dir / "data"
        data_dir.mkdir(parents=True, exist_ok=True)
        audit_path = data_dir / "audit_results.json"
        try:
            audit_path.write_text(json.dumps(audit_items, indent=2, default=str), encoding="utf-8")
        except OSError:
            pass

        total = len(audit_items)
        ok_count = sum(1 for r in audit_items if r.get("status") == "ok")

        save_history({
            "timestamp": datetime.now().isoformat(),
            "kind": "blog_audit",
            "total": total,
            "ok": ok_count,
            "fail": total - ok_count,
            "output_dir": str(self.output_dir),
            "results": audit_items,
            "collectors": ["blog_audit"],
            "fast_mode": False,
        })
